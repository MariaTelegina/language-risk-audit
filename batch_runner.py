#!/usr/bin/env python3
"""
Batch QA runner for WIT — Language Risk Audit.

Reads the 30 reviewed test cases from the "Case Input" sheet of the QA
workbook, sends each one to the /api/analyze endpoint, and writes the
results into a new "Test Runs (Auto)" sheet in a copy of the workbook —
ready for manual scoring against the Rubric sheet.

Setup:
    pip install openpyxl requests

Usage (recommended: run against your LOCAL backend, not production,
to avoid the per-visitor rate limit and to save your live Gemini quota):

    # Terminal 1
    npm run dev:api

    # Terminal 2
    python batch_runner.py \
        --input WIT_Language_Risk_Audit_30_Case_Reviewed.xlsx \
        --api-base http://127.0.0.1:5000 \
        --commit v0.3.1

Output:
    Writes WIT_Language_Risk_Audit_30_Case_Reviewed_RESULTS.xlsx next to
    the input file. The original reviewed workbook is left untouched —
    a new "Test Runs (Auto)" sheet is added to the copy.
"""

import argparse
import datetime as dt
import time

import openpyxl
import requests

CONTEXT_KEYWORDS = [
    ("customer", "Customer support"),
    ("support", "Customer support"),
    ("market", "Marketing"),
    ("workplace", "Workplace"),
    ("work", "Workplace"),
]

ALL_VARIETIES = ["American English", "Indian English", "Singapore English"]


def map_context(raw_context: str) -> str:
    """Map the free-text Context column to the app's CommunicationContext enum."""
    text = (raw_context or "").lower()
    for keyword, mapped in CONTEXT_KEYWORDS:
        if keyword in text:
            return mapped
    return "General communication"


def load_cases(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Case Input"]
    header = [cell.value for cell in ws[3]]  # row 3 holds the column headers
    col = {name: idx for idx, name in enumerate(header)}

    cases = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[col["Case ID"]]:
            continue
        cases.append(
            {
                "case_id": row[col["Case ID"]],
                "source_variety": row[col["Source variety"]],
                "original_text": row[col["Original text"]],
                "raw_context": row[col["Context"]],
            }
        )
    return cases


def run_case(api_base: str, text: str, context: str, audiences: list[str]):
    url = f"{api_base.rstrip('/')}/api/analyze"
    start = time.monotonic()
    try:
        resp = requests.post(
            url,
            json={"text": text, "selectedAudiences": audiences, "context": context},
            timeout=60,
        )
        elapsed = time.monotonic() - start
        try:
            data = resp.json()
        except ValueError:
            data = {"error": f"Non-JSON response (status {resp.status_code})"}
        return elapsed, resp.status_code, data
    except requests.RequestException as exc:
        elapsed = time.monotonic() - start
        return elapsed, None, {"error": str(exc)}


def summarize_report(data: dict) -> str:
    report = data.get("report")
    if not report:
        return f"ERROR: {data.get('error', 'no report returned')}"
    return (
        f"overallRisk={report.get('overallRisk')} | "
        f"summary={report.get('summary')} | "
        f"riskItems={len(report.get('riskItems', []))} | "
        f"rewrite={report.get('clearRewrite')}"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to the reviewed QA workbook (.xlsx)")
    parser.add_argument(
        "--api-base",
        required=True,
        help="Base URL of the backend to test, e.g. http://127.0.0.1:5000",
    )
    parser.add_argument("--commit", default="", help="App version / commit hash to record for this run")
    parser.add_argument("--output", default=None, help="Output path (defaults to <input>_RESULTS.xlsx)")
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="Seconds to wait between requests (default 1.0)",
    )
    args = parser.parse_args()

    cases = load_cases(args.input)
    print(f"Loaded {len(cases)} cases from {args.input}")

    wb = openpyxl.load_workbook(args.input)
    src_ws = wb["Test Runs"]

    # Duplicate the Test Runs sheet so the manually-scored original stays untouched
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = "Test Runs (Auto)"

    header = [cell.value for cell in new_ws[3]]
    col = {name: idx + 1 for idx, name in enumerate(header)}  # 1-indexed for openpyxl

    run_date = dt.date.today().isoformat()
    write_row = 4  # first data row, after the two header rows

    for case in cases:
        context = map_context(case["raw_context"])
        print(f"Running {case['case_id']} ({context})...")
        elapsed, status, data = run_case(args.api_base, case["original_text"], context, ALL_VARIETIES)
        summary = summarize_report(data)

        new_ws.cell(row=write_row, column=col["Case ID"], value=case["case_id"])
        new_ws.cell(row=write_row, column=col["Source variety"], value=case["source_variety"])
        new_ws.cell(row=write_row, column=col["Original text"], value=case["original_text"])
        new_ws.cell(row=write_row, column=col["Run date"], value=run_date)
        new_ws.cell(row=write_row, column=col["App version / commit"], value=args.commit)
        new_ws.cell(
            row=write_row,
            column=col["Run completed?"],
            value="Yes" if status == 200 else f"No ({status})",
        )
        new_ws.cell(row=write_row, column=col["Latency (sec)"], value=round(elapsed, 2))
        new_ws.cell(row=write_row, column=col["Actual result / summary"], value=summary)
        # Score columns (Meaning, Variety awareness, Risk identification,
        # Non-stereotyping, Recommendation, Confidence) are left blank on
        # purpose — scoring against the Rubric sheet still needs a human.

        write_row += 1
        if elapsed < args.delay:
            time.sleep(args.delay - elapsed)

    out_path = args.output or args.input.replace(".xlsx", "_RESULTS.xlsx")
    wb.save(out_path)
    print(f"\nDone. Results written to: {out_path}")
    print('Open the "Test Runs (Auto)" sheet and score each row against the Rubric sheet.')


if __name__ == "__main__":
    main()